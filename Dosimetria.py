# -*- coding: utf-8 -*-
import re
import io
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import requests
import streamlit as st
from datetime import datetime

# ============== NINOX CONFIG ==============
API_TOKEN   = "edf312a0-98b8-11f0-883e-db77626d62e5"
TEAM_ID     = "YrsYfTegptdZcHJEj"
DATABASE_ID = "ow1geqnkz00e"
BASE_URL    = "https://api.ninox.com/v1"
TABLE_WRITE_NAME = "BASE DE DATOS"

# ============== UI ==============
st.set_page_config(page_title="Microsievert — Dosimetría", page_icon="🧪", layout="wide")
st.title("🧪 Carga y Cruce de Dosis → Ninox (**BASE DE DATOS**)")

# ============== Helpers ==============
def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def hp_to_num(x) -> float:
    if x is None:
        return 0.0
    s = str(x).strip().upper()
    if s in ("", "PM", "NONE", "NAN"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def pmfmt2(v, thr: float = 0.005) -> str:
    try:
        f = float(v)
    except Exception:
        s = str(v).strip()
        return "PM" if s == "" else s
    return "PM" if f < thr else f"{f:.2f}"

def is_control_name(x: str) -> bool:
    s = strip_accents(str(x or "")).upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s.startswith("CONTROL")

def safe_cols(df: pd.DataFrame, cols: List[str]) -> List[str]:
    return [c for c in cols if c in df.columns]

# ---------- Normalización de PERIODO ----------
MES_MAP = {
    "ENE":"ENERO","FEB":"FEBRERO","MAR":"MARZO","ABR":"ABRIL","MAY":"MAYO","JUN":"JUNIO",
    "JUL":"JULIO","AGO":"AGOSTO","SEP":"SEPTIEMBRE","OCT":"OCTUBRE","NOV":"NOVIEMBRE","DIC":"DICIEMBRE",
    "JAN":"ENERO","APR":"ABRIL","AUG":"AGOSTO","DEC":"DICIEMBRE",
}
MES_NUM = {"01":"ENERO","02":"FEBRERO","03":"MARZO","04":"ABRIL","05":"MAYO","06":"JUNIO",
           "07":"JULIO","08":"AGOSTO","09":"SEPTIEMBRE","10":"OCTUBRE","11":"NOVIEMBRE","12":"DICIEMBRE"}

def _to_year4(y: str) -> str:
    y = y.strip()
    return f"20{y}" if len(y) == 2 else y

def normalizar_periodo(valor: str) -> str:
    if not valor:
        return ""
    s = strip_accents(str(valor)).upper().strip()
    s = re.sub(r"\s+", " ", s)
    m = re.match(r"^(JAN|ENE|FEB|MAR|APR|ABR|MAY|JUN|JUL|AUG|AGO|SEP|OCT|NOV|DEC|DIC)[\s\-/]*([0-9]{2,4})$", s)
    if m:
        mes = MES_MAP.get(m.group(1), m.group(1))
        anio = _to_year4(m.group(2))
        return f"{mes} {anio}"
    m = re.match(r"^([0-1][0-9])[\s\-/]*([0-9]{2,4})$", s)
    if m and m.group(1) in MES_NUM:
        mes = MES_NUM[m.group(1)]; anio = _to_year4(m.group(2))
        return f"{mes} {anio}"
    m = re.match(r"^([0-9]{4})[\s\-/]*([0-1][0-9])$", s)
    if m and m.group(2) in MES_NUM:
        mes = MES_NUM[m.group(2)]; anio = m.group(1)
        return f"{mes} {anio}"
    return s

MES_A_NUM = {"ENERO":1,"FEBRERO":2,"MARZO":3,"ABRIL":4,"MAYO":5,"JUNIO":6,"JULIO":7,"AGOSTO":8,"SEPTIEMBRE":9,"OCTUBRE":10,"NOVIEMBRE":11,"DICIEMBRE":12}
def periodo_to_date(s: str):
    if not s or not isinstance(s, str):
        return pd.NaT
    s = s.strip().upper()
    m = re.match(r"^(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+([0-9]{4})$", s)
    if not m:
        return pd.NaT
    mes = MES_A_NUM.get(m.group(1)); an = int(m.group(2))
    try:
        return pd.Timestamp(year=an, month=mes, day=1)
    except Exception:
        return pd.NaT

# ============== Ninox helpers ==============
def ninox_headers():
    return {"Authorization": f"Bearer {API_TOKEN}", "Content-Type": "application/json"}

@st.cache_data(ttl=300, show_spinner=False)
def ninox_list_tables(team_id: str, db_id: str):
    url = f"{BASE_URL}/teams/{team_id}/databases/{db_id}/tables"
    r = requests.get(url, headers=ninox_headers(), timeout=30)
    r.raise_for_status()
    return r.json()

def resolve_table_id(table_hint: str) -> str:
    hint = (table_hint or "").strip()
    if hint and " " not in hint and len(hint) <= 8:
        return hint
    for t in ninox_list_tables(TEAM_ID, DATABASE_ID):
        if str(t.get("name", "")).strip().lower() == hint.lower():
            return str(t.get("id", "")).strip()
    return hint

def ninox_insert(table_hint: str, rows: List[Dict[str, Any]], batch_size: int = 300) -> Dict[str, Any]:
    table_id = resolve_table_id(table_hint)
    url = f"{BASE_URL}/teams/{TEAM_ID}/databases/{DATABASE_ID}/tables/{table_id}/records"
    n, inserted = len(rows), 0
    if n == 0:
        return {"ok": True, "inserted": 0}
    for i in range(0, n, batch_size):
        chunk = rows[i:i+batch_size]
        r = requests.post(url, headers=ninox_headers(), json=chunk, timeout=60)
        if r.status_code != 200:
            return {"ok": False, "inserted": inserted, "error": f"{r.status_code} {r.text}"}
        inserted += len(chunk)
    return {"ok": True, "inserted": inserted}

@st.cache_data(ttl=300, show_spinner=False)
def ninox_list_records(table_hint: str, limit: int = 1000, max_pages: int = 50):
    table_id = resolve_table_id(table_hint)
    url = f"{BASE_URL}/teams/{TEAM_ID}/databases/{DATABASE_ID}/tables/{table_id}/records"
    out: List[Dict[str, Any]] = []; skip = 0
    for _ in range(max_pages):
        params = {"limit": limit, "skip": skip}
        r = requests.get(url, headers=ninox_headers(), params=params, timeout=60)
        r.raise_for_status()
        batch = r.json() or []
        if not batch: break
        out.extend(batch)
        if len(batch) < limit: break
        skip += limit
    return out

def ninox_records_to_df(records: List[Dict[str,Any]]) -> pd.DataFrame:
    if not records: return pd.DataFrame()
    rows = []
    for rec in records:
        f = rec.get("fields", {}) or {}
        rows.append({
            "PERIODO DE LECTURA": f.get("PERIODO DE LECTURA"),
            "CLIENTE": f.get("CLIENTE"),
            "CÓDIGO DE DOSÍMETRO": f.get("CÓDIGO DE DOSÍMETRO"),
            "CÓDIGO DE USUARIO": f.get("CÓDIGO DE USUARIO"),
            "NOMBRE": f.get("NOMBRE"),
            "CÉDULA": f.get("CÉDULA"),
            "FECHA DE LECTURA": f.get("FECHA DE LECTURA"),
            "TIPO DE DOSÍMETRO": f.get("TIPO DE DOSÍMETRO"),
            "Hp (10)": f.get("Hp (10)"),
            "Hp (0.07)": f.get("Hp (0.07)"),
            "Hp (3)": f.get("Hp (3)"),
        })
    df = pd.DataFrame(rows)
    if "PERIODO DE LECTURA" in df.columns:
        df["PERIODO DE LECTURA"] = df["PERIODO DE LECTURA"].astype(str).map(normalizar_periodo)
    return df

# ============== Lectores archivos ==============
def norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    def _n(x: str) -> str:
        x = strip_accents(str(x)).strip()
        x = re.sub(r"\s+", " ", x)
        return x
    out = df.copy(); out.columns = [_n(c) for c in out.columns]
    return out

def coalesce_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    got = [c for c in df.columns if c in candidates]
    return got[0] if got else None

def parse_csv_robust(upload) -> pd.DataFrame:
    raw = upload.read(); upload.seek(0)
    for sep in [";", ",", None]:
        for enc in ["utf-8-sig", "latin-1"]:
            try:
                return pd.read_csv(BytesIO(raw), sep=sep, engine="python", encoding=enc)
            except Exception:
                continue
    try:
        return pd.read_excel(BytesIO(raw))
    except Exception:
        raise

def leer_lista_codigo(upload) -> Optional[pd.DataFrame]:
    if not upload: return None
    name = upload.name.lower()
    if name.endswith((".xlsx", ".xls")):
        xls = pd.ExcelFile(upload)
        sheet = None
        for s in xls.sheet_names:
            s_norm = strip_accents(s).lower()
            if "asignar" in s_norm and "dosimet" in s_norm:
                sheet = s; break
        if sheet is None: sheet = xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet)
    else:
        df = parse_csv_robust(upload)

    df = norm_cols(df)
    c_ced   = coalesce_col(df, ["CEDULA"])
    c_user  = coalesce_col(df, ["CODIGO DE USUARIO","CODIGO_USUARIO","CODIGO USUARIO"])
    c_nom   = coalesce_col(df, ["NOMBRE","NOMBRE COMPLETO","NOMBRE_COMPLETO"])
    c_ap    = coalesce_col(df, ["APELLIDO","APELLIDOS"])
    c_cli   = coalesce_col(df, ["CLIENTE","COMPANIA"])
    c_cod   = coalesce_col(df, ["CODIGO_DOSIMETRO","CODIGO DE DOSIMETRO","CODIGO DOSIMETRO"])
    c_per   = coalesce_col(df, ["PERIODO DE LECTURA","PERIODO_DE_LECTURA","PERIODO"])
    c_tipo  = coalesce_col(df, ["TIPO DE DOSIMETRO","TIPO_DE_DOSIMETRO","TIPO DOSIMETRO"])
    c_etq   = coalesce_col(df, ["ETIQUETA"])

    out = pd.DataFrame()
    out["CÉDULA"]            = df[c_ced] if c_ced else ""
    out["CÓDIGO DE USUARIO"] = df[c_user] if c_user else ""
    if c_nom and c_ap:
        out["NOMBRE"] = (df[c_nom].astype(str).str.strip() + " " + df[c_ap].astype(str).str.strip()).str.strip()
    elif c_nom:
        out["NOMBRE"] = df[c_nom].astype(str).str.strip()
    else:
        out["NOMBRE"] = ""
    out["CLIENTE"]           = df[c_cli].astype(str).str.strip() if c_cli else ""
    out["CÓDIGO_DOSÍMETRO"]  = (df[c_cod].astype(str).str.strip().str.upper() if c_cod else "")
    out["PERIODO DE LECTURA"] = (df[c_per].astype(str).map(normalizar_periodo) if c_per else "")
    out["TIPO DE DOSÍMETRO"] = df[c_tipo].astype(str).str.strip() if c_tipo else ""
    out["ETIQUETA"]          = df[c_etq].astype(str).str.strip() if c_etq else ""

    def _is_ctrl(r):
        return is_control_name(r.get("NOMBRE","")) or is_control_name(r.get("ETIQUETA",""))
    out["_IS_CONTROL"] = out.apply(_is_ctrl, axis=1)
    return out

def leer_dosis(upload) -> Optional[pd.DataFrame]:
    if not upload: return None
    name = upload.name.lower()
    if name.endswith((".xlsx",".xls")):
        df = pd.read_excel(upload)
    else:
        df = parse_csv_robust(upload)

    cols = (df.columns.astype(str).str.strip().str.lower()
            .str.replace(" ", "", regex=False)
            .str.replace("(", "", regex=False).str.replace(")", "", regex=False)
            .str.replace(".", "", regex=False))
    df.columns = cols

    if "dosimeter" not in df.columns:
        for alt in ["dosimetro","codigo","codigodosimetro","codigo_dosimetro"]:
            if alt in df.columns:
                df.rename(columns={alt:"dosimeter"}, inplace=True); break

    for cands, dest in [ (["hp10dosecorr","hp10dose","hp10"], "hp10dose"),
                         (["hp007dosecorr","hp007dose","hp007"], "hp0.07dose"),
                         (["hp3dosecorr","hp3dose","hp3"], "hp3dose") ]:
        for c in cands:
            if c in df.columns: df.rename(columns={c:dest}, inplace=True); break
        if dest not in df.columns: df[dest] = 0.0
        else: df[dest] = pd.to_numeric(df[dest], errors="coerce").fillna(0.0)

    if "dosimeter" in df.columns:
        df["dosimeter"] = df["dosimeter"].astype(str).str.strip().str.upper()
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    return df

# ============== Construcción de registros ==============
def construir_registros(df_lista: pd.DataFrame, df_dosis: pd.DataFrame, periodos: List[str]) -> pd.DataFrame:
    df_l = df_lista.copy()
    df_l["PERIODO DE LECTURA"] = df_l["PERIODO DE LECTURA"].astype(str).str.strip().str.upper()
    df_l["CÓDIGO_DOSÍMETRO"]   = df_l["CÓDIGO_DOSÍMETRO"].astype(str).str.strip().str.upper()

    selected = [p.strip().upper() for p in periodos if str(p).strip()]
    if selected:
        df_l = df_l[df_l["PERIODO DE LECTURA"].isin(selected)]

    idx = df_dosis.set_index("dosimeter") if "dosimeter" in df_dosis.columns else pd.DataFrame().set_index(pd.Index([]))
    registros: List[Dict[str, Any]] = []
    df_l = pd.concat([df_l[df_l["_IS_CONTROL"]], df_l[~df_l["_IS_CONTROL"]]], ignore_index=True)

    for _, r in df_l.iterrows():
        cod = str(r["CÓDIGO_DOSÍMETRO"]).strip().upper()
        if not cod or cod == "NAN": continue
        if cod not in idx.index:    continue
        d = idx.loc[cod]
        if isinstance(d, pd.DataFrame):
            d = d.sort_values(by="timestamp").iloc[-1]
        ts = d.get("timestamp", pd.NaT)
        try:
            fecha_str = pd.to_datetime(ts).strftime("%d/%m/%Y %H:%M") if pd.notna(ts) else ""
        except Exception:
            fecha_str = ""

        nombre = str(r.get("NOMBRE","")).strip()
        if bool(r["_IS_CONTROL"]) and (not nombre):
            nombre = "CONTROL"

        registros.append({
            "PERIODO DE LECTURA": r["PERIODO DE LECTURA"],
            "CLIENTE": r.get("CLIENTE",""),
            "CÓDIGO DE DOSÍMETRO": cod,
            "CÓDIGO DE USUARIO": r.get("CÓDIGO DE USUARIO",""),
            "NOMBRE": nombre,
            "CÉDULA": r.get("CÉDULA",""),
            "FECHA DE LECTURA": fecha_str,
            "TIPO DE DOSÍMETRO": r.get("TIPO DE DOSÍMETRO","") or "CE",
            "Hp (10)":  float(d.get("hp10dose", 0.0) or 0.0),
            "Hp (0.07)":float(d.get("hp0.07dose", 0.0) or 0.0),
            "Hp (3)":   float(d.get("hp3dose", 0.0) or 0.0),
            "_IS_CONTROL": bool(r["_IS_CONTROL"])
        })

    df_final = pd.DataFrame(registros)
    if not df_final.empty:
        df_final = df_final.sort_values(["_IS_CONTROL","NOMBRE","CÉDULA"], ascending=[False, True, True]).reset_index(drop=True)
    return df_final

# ============== Resta de CONTROL + Formato ==============
def aplicar_resta_control_y_formato(df_final: pd.DataFrame, umbral_pm: float = 0.005, manual_ctrl: Optional[float] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if df_final is None or df_final.empty:
        return df_final, df_final

    df = df_final.copy()
    for h in ["Hp (10)", "Hp (0.07)", "Hp (3)"]:
        if h not in df.columns:
            df[h] = 0.0
        df[h] = pd.to_numeric(df[h], errors="coerce").fillna(0.0)
    if "PERIODO DE LECTURA" in df.columns:
        df["PERIODO DE LECTURA"] = df["PERIODO DE LECTURA"].astype(str).map(normalizar_periodo)

    is_ctrl = df["NOMBRE"].apply(is_control_name)
    df_ctrl = df[is_ctrl].copy()
    df_per  = df[~is_ctrl].copy()

    ctrl_means = pd.DataFrame(columns=["PERIODO DE LECTURA","Hp10_CTRL","Hp007_CTRL","Hp3_CTRL"])
    if not df_ctrl.empty:
        ctrl_means = df_ctrl.groupby("PERIODO DE LECTURA", as_index=False).agg({
            "Hp (10)":"mean","Hp (0.07)":"mean","Hp (3)":"mean"
        }).rename(columns={"Hp (10)":"Hp10_CTRL","Hp (0.07)":"Hp007_CTRL","Hp (3)":"Hp3_CTRL"})

    out = df_per.copy()
    if not ctrl_means.empty:
        out = out.merge(ctrl_means, on="PERIODO DE LECTURA", how="left")
        out["Hp10_CTRL"]  = out["Hp10_CTRL"].fillna(0.0)
        out["Hp007_CTRL"] = out["Hp007_CTRL"].fillna(0.0)
        out["Hp3_CTRL"]   = out["Hp3_CTRL"].fillna(0.0)
        out["_Hp10_NUM"]  = (out["Hp (10)"]   - out["Hp10_CTRL"]).clip(lower=0.0)
        out["_Hp007_NUM"] = (out["Hp (0.07)"] - out["Hp007_CTRL"]).clip(lower=0.0)
        out["_Hp3_NUM"]   = (out["Hp (3)"]    - out["Hp3_CTRL"]).clip(lower=0.0)
    else:
        if manual_ctrl is not None and float(manual_ctrl) > 0:
            cval = float(manual_ctrl)
            out["_Hp10_NUM"]  = (out["Hp (10)"]   - cval).clip(lower=0.0)
            out["_Hp007_NUM"] = (out["Hp (0.07)"] - cval).clip(lower=0.0)
            out["_Hp3_NUM"]   = (out["Hp (3)"]    - cval).clip(lower=0.0)
        else:
            out["_Hp10_NUM"]  = out["Hp (10)"]
            out["_Hp007_NUM"] = out["Hp (0.07)"]
            out["_Hp3_NUM"]   = out["Hp (3)"]

    out_view = out.copy()
    out_view["Hp (10)"]   = out_view["_Hp10_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
    out_view["Hp (0.07)"] = out_view["_Hp007_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
    out_view["Hp (3)"]    = out_view["_Hp3_NUM"].map(lambda v: pmfmt2(v, umbral_pm))

    df_ctrl_view = pd.DataFrame()
    if not df_ctrl.empty:
        df_ctrl_view = df_ctrl.merge(ctrl_means, on="PERIODO DE LECTURA", how="left")
        for c in ["Hp10_CTRL","Hp007_CTRL","Hp3_CTRL"]:
            df_ctrl_view[c] = df_ctrl_view[c].fillna(0.0)
        df_ctrl_view["_Hp10_NUM"]  = (df_ctrl_view["Hp (10)"]   - df_ctrl_view["Hp10_CTRL"]).clip(lower=0.0)
        df_ctrl_view["_Hp007_NUM"] = (df_ctrl_view["Hp (0.07)"] - df_ctrl_view["Hp007_CTRL"]).clip(lower=0.0)
        df_ctrl_view["_Hp3_NUM"]   = (df_ctrl_view["Hp (3)"]    - df_ctrl_view["Hp3_CTRL"]).clip(lower=0.0)
        df_ctrl_view["Hp (10)"]    = df_ctrl_view["_Hp10_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
        df_ctrl_view["Hp (0.07)"]  = df_ctrl_view["_Hp007_NUM"].map(lambda v: pmfmt2(v, umbral_pm))
        df_ctrl_view["Hp (3)"]     = df_ctrl_view["_Hp3_NUM"].map(lambda v: pmfmt2(v, umbral_pm))

    df_vista = pd.concat([df_ctrl_view, out_view], ignore_index=True, sort=False)
    if not df_vista.empty:
        df_vista["__is_control__"] = df_vista["NOMBRE"].apply(is_control_name)
        df_vista = df_vista.sort_values(by=["__is_control__","NOMBRE","CÉDULA"], ascending=[False, True, True]).drop(columns=["__is_control__"])

    df_num = out[["_Hp10_NUM","_Hp007_NUM","_Hp3_NUM","PERIODO DE LECTURA","CLIENTE",
                  "CÓDIGO DE USUARIO","CÓDIGO DE DOSÍMETRO","NOMBRE","CÉDULA",
                  "TIPO DE DOSÍMETRO","FECHA DE LECTURA"]].copy()
    return df_vista, df_num

# ============== REPORTE ÚNICO (CONTROL primero) ==============
def construir_reporte_unico(df_vista: pd.DataFrame, df_num: pd.DataFrame, umbral_pm: float = 0.005, agrupar_control_por: str = "CLIENTE") -> pd.DataFrame:
    if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
        return pd.DataFrame()

    # Personas
    personas_num = df_num[~df_num["NOMBRE"].apply(is_control_name)].copy()
    if not personas_num.empty:
        per_anual = personas_num.groupby("CÓDIGO DE USUARIO", as_index=False).agg({
            "CLIENTE":"last","NOMBRE":"last","CÉDULA":"last","CÓDIGO DE DOSÍMETRO":"last",
            "_Hp10_NUM":"sum","_Hp007_NUM":"sum","_Hp3_NUM":"sum"
        }).rename(columns={"_Hp10_NUM":"Hp (10) ANUAL","_Hp007_NUM":"Hp (0.07) ANUAL","_Hp3_NUM":"Hp (3) ANUAL"})
        personas_num["__fecha__"] = personas_num["PERIODO DE LECTURA"].map(periodo_to_date)
        idx_last = personas_num.groupby("CÓDIGO DE USUARIO")["__fecha__"].idxmax()
        per_last = personas_num.loc[idx_last, safe_cols(personas_num, [
            "CÓDIGO DE USUARIO","PERIODO DE LECTURA","_Hp10_NUM","_Hp007_NUM","_Hp3_NUM",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO"
        ])].rename(columns={"_Hp10_NUM":"Hp (10)","_Hp007_NUM":"Hp (0.07)","_Hp3_NUM":"Hp (3)"})
        per_view = per_anual.merge(per_last, on="CÓDIGO DE USUARIO", how="left")
        for c in safe_cols(per_view, ["Hp (10)","Hp (0.07)","Hp (3)","Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL"]):
            per_view[c] = per_view[c].map(lambda v: pmfmt2(v, umbral_pm))
        per_view["Hp (10) DE POR VIDA"]   = per_view["Hp (10) ANUAL"]
        per_view["Hp (0.07) DE POR VIDA"] = per_view["Hp (0.07) ANUAL"]
        per_view["Hp (3) DE POR VIDA"]    = per_view["Hp (3) ANUAL"]
        personas_final = per_view[safe_cols(per_view, [
            "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE","CÉDULA",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)",
            "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA"
        ])]
    else:
        personas_final = pd.DataFrame(columns=[
            "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE","CÉDULA",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)",
            "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA"
        ])

    # Control (una fila por CLIENTE)
    control_v = df_vista[df_vista["NOMBRE"].apply(is_control_name)].copy()
    if not control_v.empty:
        for h in safe_cols(control_v, ["Hp (10)","Hp (0.07)","Hp (3)"]):
            control_v[h] = control_v[h].apply(hp_to_num)
        agr = agrupar_control_por if agrupar_control_por in control_v.columns else None
        if agr is None:
            control_v["__grupo__"] = "GLOBAL"; agr = "__grupo__"

        ctrl_anual = control_v.groupby(agr, as_index=False).agg({
            "CLIENTE":"last","Hp (10)":"sum","Hp (0.07)":"sum","Hp (3)":"sum"
        }).rename(columns={"Hp (10)":"Hp (10) ANUAL","Hp (0.07)":"Hp (0.07) ANUAL","Hp (3)":"Hp (3) ANUAL"})
        tmp = control_v.copy()
        tmp["__fecha__"] = tmp["PERIODO DE LECTURA"].map(periodo_to_date)
        idx_last_c = tmp.groupby(agr)["__fecha__"].idxmax()
        last_vals = tmp.loc[idx_last_c, safe_cols(tmp, [
            agr,"PERIODO DE LECTURA","Hp (10)","Hp (0.07)","Hp (3)",
            "CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","CÉDULA","FECHA DE LECTURA","TIPO DE DOSÍMETRO"
        ])]
        ctrl_view = ctrl_anual.merge(last_vals, on=agr, how="left")
        ctrl_view["NOMBRE"] = "CONTROL"

        def _fill_usercode(row):
            cu = str(row.get("CÓDIGO DE USUARIO","") or "").strip()
            return cu if cu else str(row.get("CÓDIGO DE DOSÍMETRO","") or "").strip()
        ctrl_view["CÓDIGO DE USUARIO"] = ctrl_view.apply(_fill_usercode, axis=1)

        for c in safe_cols(ctrl_view, ["Hp (10)","Hp (0.07)","Hp (3)","Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL"]):
            ctrl_view[c] = ctrl_view[c].map(lambda v: pmfmt2(v, umbral_pm))
        ctrl_view["Hp (10) DE POR VIDA"]   = ctrl_view["Hp (10) ANUAL"]
        ctrl_view["Hp (0.07) DE POR VIDA"] = ctrl_view["Hp (0.07) ANUAL"]
        ctrl_view["Hp (3) DE POR VIDA"]    = ctrl_view["Hp (3) ANUAL"]

        ctrl_final = ctrl_view[safe_cols(ctrl_view, [
            "PERIODO DE LECTURA","CLIENTE","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE","CÉDULA",
            "FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)",
            "Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA","Hp (3) DE POR VIDA"
        ])]
    else:
        ctrl_final = pd.DataFrame(columns=personas_final.columns)

    reporte = pd.concat([ctrl_final, personas_final], ignore_index=True)
    if not reporte.empty:
        reporte["__is_control__"] = reporte["NOMBRE"].apply(is_control_name)
        # Ordenar por CÓDIGO DE DOSÍMETRO (como pediste) y dentro por usuario
        reporte = reporte.sort_values(
            by=["__is_control__","CÓDIGO DE DOSÍMETRO","CÓDIGO DE USUARIO","NOMBRE"],
            ascending=[False, True, True, True]
        ).drop(columns=["__is_control__"])
    return reporte

# ============== Excel con diseño (como el ejemplo) ==============
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

THIN = Side(color="000000", style="thin")
BORD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GREY = PatternFill("solid", fgColor="DDDDDD")
LIGHT = PatternFill("solid", fgColor="F5F5F5")

def _box(ws, r0, c0, r1, c1, header=False, center=False, fill=None, bold=False):
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(r, c)
            cell.border = BORD
            if fill: cell.fill = fill
            if header:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif center:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if bold: cell.font = Font(bold=True)

def build_excel_like_example(df_reporte: pd.DataFrame, fecha_emision: str, cliente: str, codigo_reporte: str, logo_bytes: Optional[bytes] = None) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "REPORTE"
    widths = {1:14,2:14,3:26,4:16,5:20,6:14,7:16,8:10,9:10,10:10,11:10,12:10,13:10,14:10,15:10}
    for col, w in widths.items(): ws.column_dimensions[chr(64+col)].width = w

    row = 1
    if logo_bytes:
        img = XLImage(BytesIO(logo_bytes)); img.width=240; img.height=88
        ws.add_image(img, "A1"); row = 7
    else:
        row = 3

    ws.cell(row,1,"MICROSIEVERT, S.A.").font=Font(bold=True); row+=1
    ws.cell(row,1,"PH Conardo"); row+=1
    ws.cell(row,1,"Calle 41 Este, Panamá"); row+=1
    ws.cell(row,1,"PANAMÁ"); row+=2

    ws.merge_cells(start_row=row-3, start_column=10, end_row=row-3, end_column=12)
    ws.cell(row-3,10,"Fecha de emisión").alignment=Alignment(horizontal="center")
    _box(ws,row-3,10,row-3,12,header=True,fill=GREY)
    ws.merge_cells(start_row=row-2, start_column=10, end_row=row-2, end_column=12)
    ws.cell(row-2,10,fecha_emision).alignment=Alignment(horizontal="center")
    _box(ws,row-2,10,row-2,12,center=True)

    ws.merge_cells(start_row=row-3, start_column=13, end_row=row-3, end_column=15)
    ws.cell(row-3,13,"Cliente").alignment=Alignment(horizontal="center")
    _box(ws,row-3,13,row-3,15,header=True,fill=GREY)
    ws.merge_cells(start_row=row-2, start_column=13, end_row=row-2, end_column=15)
    ws.cell(row-2,13,cliente).alignment=Alignment(horizontal="center")
    _box(ws,row-2,13,row-2,15,center=True)

    ws.merge_cells(start_row=row-1, start_column=13, end_row=row-1, end_column=15)
    ws.cell(row-1,13,"Código").alignment=Alignment(horizontal="center")
    _box(ws,row-1,13,row-1,15,header=True,fill=GREY)
    ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=15)
    ws.cell(row,13,codigo_reporte).alignment=Alignment(horizontal="center")
    _box(ws,row,13,row,15,center=True)
    row += 2

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
    ws.cell(row,6,"REPORTE DE DOSIMETRÍA").font=Font(bold=True)
    ws.cell(row,6).alignment=Alignment(horizontal="center")
    row += 2

    # Cabecera agrupada
    cab1 = [("DATOS DEL USUARIO Y DE LA LECTURA DOSIMÉTRICA ",1,6),
            ("DOSIS ACTUAL (mSv) ",7,9),
            ("DOSIS ANUAL  (mSv) ",10,12),("DOSIS DE POR VIDA (mSv)",13,15)
    ]
   
    for txt,c0,c1 in cab1:
            ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c1)
            ws.cell(row,c0,txt)
    _box(ws,row,1,row,15,header=True,fill=LIGHT)
    row += 1

    # Subcabeceras
    ws.cell(row,1,"PERIODO DE LECTURA")
    ws.cell(row,2,"CÓDIGO DE USUARIO")
    ws.cell(row,3,"NOMBRE")
    ws.cell(row,4,"CÉDULA")
    ws.cell(row,5,"FECHA DE LECTURA")
    ws.cell(row,6,"TIPO DE DOSÍMETRO")
    ws.cell(row,7,"Hp(10)"); ws.cell(row,8,"Hp(0.07)"); ws.cell(row,9,"Hp(3)")
    ws.cell(row,10,"Hp(10)"); ws.cell(row,11,"Hp(0.07)"); ws.cell(row,12,"Hp(3)")
    ws.cell(row,13,"Hp(10)"); ws.cell(row,14,"Hp(0.07)"); ws.cell(row,15,"Hp(3)")
    _box(ws,row,1,row,15,header=True,fill=GREY)

    # Datos
    start_data = row + 1
    cols = ["PERIODO DE LECTURA","CÓDIGO DE USUARIO","NOMBRE","CÉDULA","FECHA DE LECTURA","TIPO DE DOSÍMETRO",
            "Hp (10)","Hp (0.07)","Hp (3)","Hp (10) ANUAL","Hp (0.07) ANUAL","Hp (3) ANUAL",
            "Hp (10) DE POR VIDA","Hp (0.07) DE POR VIDA", "Hp (3) DE POR VIDA" ]
    df_to_write = df_reporte[[c for c in cols if c in df_reporte.columns]].copy()
    # Encajar en columnas 1..15
    for rr in dataframe_to_rows(df_to_write, index=False, header=False):
        for j, v in enumerate(rr, start=1):
            ws.cell(start_data, j, v)
        start_data += 1
    end_data = start_data - 1
    _box(ws, row-1, 1, end_data, 15)  # toda la tabla

    # ====== INFORMACIÓN (debajo de la tabla)
    row = end_data + 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    ws.cell(row,1,"INFORMACIÓN DEL REPORTE DE DOSIMETRÍA").font=Font(bold=True)
    ws.cell(row,1).alignment=Alignment(horizontal="center")
    _box(ws,row,1,row,15,header=True,fill=LIGHT)
    row += 2

    for txt in [
        "– Periodo de lectura: periodo de uso del dosímetro personal.",
        "– Fecha de lectura: corresponde a la fecha en que fue realizada la lectura del dosímetro.",
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        ws.cell(row,1,txt); _box(ws,row,1,row,15); row += 2

    # Tipo de dosímetro (izq) y límites (der)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row,1,"– Tipo de dosímetro:"); _box(ws,row,1,row,5); row += 1

    r0 = row
    for i, texto in enumerate(["CE = Cuerpo Entero","A = Anillo","B = Brazalete","CR = Cristalino"]):
        ws.merge_cells(start_row=r0+i, start_column=2, end_row=r0+i, end_column=5)
        ws.cell(r0+i,2,texto); _box(ws,r0+i,2,r0+i,5,center=True)
    rL0 = row
    ws.merge_cells(start_row=rL0, start_column=8, end_row=rL0, end_column=15)
    ws.cell(rL0,8,"LÍMITES ANUALES DE EXPOSICIÓN A RADIACIONES").alignment=Alignment(horizontal="center")
    _box(ws,rL0,8,rL0,15,header=True,fill=GREY)
    limites = [("Cuerpo Entero","20mSv/año"),("Cristalino","150 mSv/año"),
               ("Extremidades y piel","500 mSv/año"),("Fetal","1 mSv/periodo de gestación"),
               ("Público","1 mSv/año")]
    rr = rL0 + 1
    for nom,val in limites:
        ws.merge_cells(start_row=rr, start_column=8, end_row=rr, end_column=11)
        ws.cell(rr,8,nom); _box(ws,rr,8,rr,11)
        ws.merge_cells(start_row=rr, start_column=12, end_row=rr, end_column=15)
        ws.cell(rr,12,val).alignment=Alignment(horizontal="center")
        _box(ws,rr,12,rr,15,center=True); rr += 1
    row = max(r0+4, rr) + 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    ws.cell(row,1,"– DATOS DEL PARTICIPANTE:").font=Font(bold=True)
    _box(ws,row,1,row,15,fill=LIGHT); row += 1
    bullets = [
        "- Código de usuario: Número único asignado al usuario por Microsievert, S.A.",
        "- Nombre: Persona a la cual se le asigna el dosímetro personal.",
        "- Cédula: Número del documento de identidad personal del usuario.",
        "- Fecha de nacimiento: Registro de la fecha de nacimiento del usuario.",
    ]
    for b in bullets:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        ws.cell(row,1,b); _box(ws,row,1,row,15); row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
    ws.cell(row,1,"– DOSIS EN MILISIEVERT:").font=Font(bold=True)
    _box(ws,row,1,row,15,fill=LIGHT); row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5); ws.cell(row,2,"Nombre")
    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=13); ws.cell(row,6,"Definición")
    ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=15); ws.cell(row,14,"Unidad")
    _box(ws,row,2,row,15,header=True,fill=GREY); row += 1

    filas_dosis = [
        ("Hp(10) = Dosis efectiva","Es la dosis equivalente en tejido blando, J·kg-1 ó Sv a una profundidad de 10 mm, bajo determinado punto","mSv"),
        ("Hp(0.07) = Dosis equivalente superficial","Es la dosis equivalente en tejido blando, J·kg-1 ó Sv a una profundidad de 0,07 mm, bajo determinado punto","mSv"),
        ("Hp(3) = Dosis equivalente a cristalino","Es la dosis equivalente en tejido blando, J·kg-1 ó Sv a una profundidad de 3 mm, bajo determinado punto del","mSv"),
    ]
    for nom,defin,uni in filas_dosis:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5); ws.cell(row,2,nom); _box(ws,row,2,row,5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=13); ws.cell(row,6,defin); _box(ws,row,6,row,13)
        ws.merge_cells(start_row=row, start_column=14, end_row=row, end_column=15); ws.cell(row,14,uni).alignment=Alignment(horizontal="center"); _box(ws,row,14,row,15,center=True)
        row += 1

    row += 1
    textos = [
        "LECTURAS DE ANILLO: las lecturas del dosímetro de anillo son registradas como una dosis equivalente superficial Hp(0.07)",
        "Los resultados de las dosis individuales de radiación son reportados para diferentes periodos de tiempo:",
        "DOSÍMETRO DE CONTROL: incluido en cada paquete entregado para monitorear la exposición a la radiación recibida durante el tránsito y almacenamiento. Este dosímetro",
        "POR DEBAJO DEL MÍNIMO REPORTABLE: es la dosis por debajo de la cantidad mínima reportada para el periodo de uso y son registradas como \"PM\".",
    ]
    for t in textos:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=15)
        ws.cell(row,1,t); _box(ws,row,1,row,15); row += 1

    labels = [("DOSIS ACTUAL","Es el correspondiente de dosis acumulada durante el periodo de lectura definido."),
              ("DOSIS ANUAL","Es el correspondiente de dosis acumulada desde el inicio del año hasta la fecha."),
              ("DOSIS DE POR VIDA","Es el correspondiente de DOSIS acumulada desde el inicio del servicio dosimétrico hasta la fecha.")]
    for lab, txt in labels:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5); ws.cell(row,2,lab); _box(ws,row,2,row,5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=15); ws.cell(row,6,txt); _box(ws,row,6,row,15)
        row += 1

    bio = BytesIO(); wb.save(bio); return bio.getvalue()

# ============== UI: Tabs ==============
tab1, tab2 = st.tabs(["1) Cargar y Subir a Ninox", "2) Reporte Final (sumas)"])

# -------- TAB 1
with tab1:
    st.subheader("1) Cargar LISTA DE CÓDIGO")
    upl_lista = st.file_uploader("Sube la LISTA DE CÓDIGO (CSV / XLS / XLSX)", type=["csv","xls","xlsx"], key="upl_lista")
    df_lista = leer_lista_codigo(upl_lista) if upl_lista else None
    if df_lista is not None and not df_lista.empty:
        st.success(f"LISTA cargada: {len(df_lista)} filas")
        st.dataframe(df_lista.head(20), use_container_width=True)
    else:
        st.info("LISTA vacía o sin datos")

    st.subheader("2) Subir Archivo de Dosis")
    upl_dosis = st.file_uploader("Selecciona CSV/XLS/XLSX (dosis)", type=["csv","xls","xlsx"], key="upl_dosis")
    df_dosis = leer_dosis(upl_dosis) if upl_dosis else None
    if df_dosis is not None and not df_dosis.empty:
        st.success(f"Dosis cargadas: {len(df_dosis)} fila(s)")
        st.dataframe(df_dosis.head(15), use_container_width=True)

    per_options = sorted(df_lista["PERIODO DE LECTURA"].dropna().astype(str).str.upper().unique().tolist()) if df_lista is not None else []
    periodos_sel = st.multiselect("Filtrar por PERIODO DE LECTURA (elige uno o varios; vacío = TODOS)", per_options, default=[])

    with st.expander("⚙️ Opcional: Control manual si NO existe CONTROL en el periodo"):
        use_manual_ctrl = st.checkbox("Activar control manual", value=False)
        manual_ctrl_val = st.number_input("Valor de control manual a restar (Hp10, Hp0.07, Hp3)", min_value=0.0, step=0.001, format="%.3f", value=0.000)

    subir_pm_como_texto = st.checkbox("Guardar 'PM' como texto en Ninox (si desmarcas, sube None en PM)", value=True)

    # >>> NEW: helper para llave y set de existentes
    def _mk_key(row: Dict[str, Any]) -> Tuple[str,str,str,str]:
        return (
            str(row.get("PERIODO DE LECTURA","")).strip().upper(),
            str(row.get("CÓDIGO DE DOSÍMETRO","")).strip().upper(),
            str(row.get("CÓDIGO DE USUARIO","")).strip(),
            strip_accents(str(row.get("NOMBRE","")).strip().upper()),
        )

    @st.cache_data(ttl=120, show_spinner=False)
    def _fetch_existing_keys() -> set:
        recs = ninox_list_records(TABLE_WRITE_NAME, limit=1000)
        df = ninox_records_to_df(recs)
        if df.empty:
            return set()
        keys = set()
        for _, r in df.iterrows():
            keys.add(_mk_key(r))
        return keys

    if st.button("✅ Procesar y Previsualizar", type="primary"):
        if df_lista is None or df_lista.empty:
            st.error("Primero sube la LISTA DE CÓDIGO.")
        elif df_dosis is None or df_dosis.empty:
            st.error("Sube el archivo de dosis.")
        elif "dosimeter" not in df_dosis.columns:
            st.error("El archivo de dosis debe incluir la columna 'dosimeter'.")
        else:
            df_final_raw = construir_registros(df_lista, df_dosis, periodos_sel)
            if df_final_raw.empty:
                with st.expander("Debug de coincidencias (no se encontraron)"):
                    st.write({
                        "dosimeter únicos en dosis": sorted(df_dosis["dosimeter"].dropna().unique().tolist()) if "dosimeter" in df_dosis.columns else [],
                        "CÓDIGO_DOSÍMETRO únicos en LISTA (según filtro)": sorted(df_lista["CÓDIGO_DOSÍMETRO"].dropna().unique().tolist()) if "CÓDIGO_DOSÍMETRO" in df_lista.columns else []
                    })
                st.warning("⚠️ No hay coincidencias **CÓDIGO_DOSÍMETRO** ↔ **dosimeter** (revisa periodos/códigos).")
            else:
                df_vista, df_num_corr = aplicar_resta_control_y_formato(
                    df_final_raw, umbral_pm=0.005,
                    manual_ctrl=(manual_ctrl_val if use_manual_ctrl else None)
                )
                st.session_state.df_final_vista = df_vista.drop(columns=["_IS_CONTROL"], errors="ignore")
                st.session_state.df_final_num   = df_num_corr
                st.success(f"¡Listo! Registros generados (corregidos): {len(st.session_state.df_final_vista)}")
                st.dataframe(st.session_state.df_final_vista, use_container_width=True)
                st.caption(f"Controles detectados: {(st.session_state.df_final_vista['NOMBRE'].apply(is_control_name)).sum()}")

    st.markdown("---")
    st.subheader("3) Subir a Ninox")

    def _to_str(v):
        if pd.isna(v): return ""
        if isinstance(v, (pd.Timestamp, )):
            return v.strftime("%Y-%m-%d %H:%M:%S")
        return str(v)

    # >>> CHANGED: ahora acepta flag is_control
    def _hp_value_for_upload(v, as_text_pm=True, is_control=False):
        """
        - Personas: si es 'PM' -> 'PM' (cuando as_text_pm=True), o None si as_text_pm=False.
        - CONTROL: nunca 'PM'; si viene 'PM' lo convertimos a 0.00 (subir como número o string según as_text_pm).
        """
        sv = None if v is None else str(v).strip().upper()
        if is_control:
            # CONTROL: forzar número
            try:
                num = float(v)
            except Exception:
                num = 0.0 if sv == "PM" or sv in ("", "NONE", "NAN") else hp_to_num(v)
            return f"{num:.2f}" if as_text_pm else num
        else:
            # Personas
            if isinstance(v, str) and sv == "PM":
                return "PM" if as_text_pm else None
            try:
                num = float(v)
            except Exception:
                return v if v is not None else None
            return f"{num:.2f}" if as_text_pm else num

    # >>> NEW: botón de actualización (solo nuevos)
    if st.button("🔁 Actualizar en Ninox (solo NUEVOS)"):
        df_vista = st.session_state.get("df_final_vista")
        df_num   = st.session_state.get("df_final_num")
        if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
            st.error("No hay datos procesados. Pulsa 'Procesar y Previsualizar' primero.")
        else:
            df_para_subir = construir_reporte_unico(df_vista, df_num)  # consolidado
            if df_para_subir.empty:
                st.error("Nada para subir después de consolidar.")
            else:
                existing = _fetch_existing_keys()
                rows = []
                nuevos = 0
                for _, rowx in df_para_subir.iterrows():
                    key = _mk_key(rowx)
                    if key in existing:
                        continue  # ya existe -> no subir
                    is_ctrl = is_control_name(rowx.get("NOMBRE",""))
                    fields = {
                        "PERIODO DE LECTURA": _to_str(rowx.get("PERIODO DE LECTURA","")),
                        "CLIENTE": _to_str(rowx.get("CLIENTE","")),
                        "CÓDIGO DE DOSÍMETRO": _to_str(rowx.get("CÓDIGO DE DOSÍMETRO","")),
                        "CÓDIGO DE USUARIO": _to_str(rowx.get("CÓDIGO DE USUARIO","")),
                        "NOMBRE": _to_str(rowx.get("NOMBRE","")),
                        "CÉDULA": _to_str(rowx.get("CÉDULA","")),
                        "FECHA DE LECTURA": _to_str(rowx.get("FECHA DE LECTURA","")),
                        "TIPO DE DOSÍMETRO": _to_str(rowx.get("TIPO DE DOSÍMETRO","") or "CE"),
                        "Hp (10)": _hp_value_for_upload(rowx.get("Hp (10)"), as_text_pm=subir_pm_como_texto, is_control=is_ctrl),
                        "Hp (0.07)": _hp_value_for_upload(rowx.get("Hp (0.07)"), as_text_pm=subir_pm_como_texto, is_control=is_ctrl),
                        "Hp (3)": _hp_value_for_upload(rowx.get("Hp (3)"), as_text_pm=subir_pm_como_texto, is_control=is_ctrl),
                    }
                    rows.append({"fields": fields})
                    nuevos += 1
                if nuevos == 0:
                    st.info("No hay registros NUEVOS que subir. Todo ya existe en Ninox con la llave definida.")
                else:
                    with st.spinner(f"Subiendo {nuevos} registro(s) nuevo(s) a Ninox..."):
                        res = ninox_insert(TABLE_WRITE_NAME, rows, batch_size=300)
                    if res.get("ok"):
                        st.success(f"✅ Subido a Ninox: {res.get('inserted', 0)} registro(s) NUEVO(s).")
                        st.toast("¡Actualización completada!", icon="✅")
                        _fetch_existing_keys.clear()  # invalidar caché
                    else:
                        st.error(f"❌ Error al subir: {res.get('error')}")

    # Botón original (sube TODO lo consolidado)
    if st.button("⬆️ Subir a Ninox (BASE DE DATOS)"):
        df_vista = st.session_state.get("df_final_vista")
        df_num   = st.session_state.get("df_final_num")
        if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
            st.error("No hay datos procesados. Pulsa 'Procesar y Previsualizar' primero.")
        else:
            df_para_subir = construir_reporte_unico(df_vista, df_num)  # consolidado
            if df_para_subir.empty:
                st.error("Nada para subir después de consolidar.")
            else:
                rows = []
                for _, rowx in df_para_subir.iterrows():
                    is_ctrl = is_control_name(rowx.get("NOMBRE",""))
                    fields = {
                        "PERIODO DE LECTURA": _to_str(rowx.get("PERIODO DE LECTURA","")),
                        "CLIENTE": _to_str(rowx.get("CLIENTE","")),
                        "CÓDIGO DE DOSÍMETRO": _to_str(rowx.get("CÓDIGO DE DOSÍMETRO","")),
                        "CÓDIGO DE USUARIO": _to_str(rowx.get("CÓDIGO DE USUARIO","")),
                        "NOMBRE": _to_str(rowx.get("NOMBRE","")),
                        "CÉDULA": _to_str(rowx.get("CÉDULA","")),
                        "FECHA DE LECTURA": _to_str(rowx.get("FECHA DE LECTURA","")),
                        "TIPO DE DOSÍMETRO": _to_str(rowx.get("TIPO DE DOSÍMETRO","") or "CE"),
                        # >>> CHANGED: CONTROL no sube 'PM' jamás; personas mantienen 'PM' si aplica
                        "Hp (10)": _hp_value_for_upload(rowx.get("Hp (10)"), as_text_pm=subir_pm_como_texto, is_control=is_ctrl),
                        "Hp (0.07)": _hp_value_for_upload(rowx.get("Hp (0.07)"), as_text_pm=subir_pm_como_texto, is_control=is_ctrl),
                        "Hp (3)": _hp_value_for_upload(rowx.get("Hp (3)"), as_text_pm=subir_pm_como_texto, is_control=is_ctrl),
                    }
                    rows.append({"fields": fields})
                with st.spinner("Subiendo a Ninox..."):
                    res = ninox_insert(TABLE_WRITE_NAME, rows, batch_size=300)
                if res.get("ok"):
                    st.success(f"✅ Subido a Ninox: {res.get('inserted', 0)} registro(s).")
                    st.toast("¡Datos enviados a Ninox!", icon="✅")
                else:
                    st.error(f"❌ Error al subir: {res.get('error')}")

# -------- TAB 2
with tab2:
    st.subheader("📊 Reporte Final (CONTROL primero y luego PERSONAS)")
    fuente = st.radio("Fuente de datos para el reporte:", [
        "Usar datos procesados en esta sesión",
        "Leer directamente de Ninox (tabla BASE DE DATOS)",
    ], index=0)

    # >>> NEW: Nombre de archivo deseado
    nombre_archivo_base = st.text_input("Nombre de archivo para las descargas (sin extensión)", value="Reporte_Final")  # se usa para CSV/Excel

    # Datos del encabezado del reporte
    codigo_reporte_ui = st.text_input("Código del reporte (opcional)", value="SIN-CÓDIGO")
    fecha_emision_ui = st.date_input("Fecha de emisión", value=pd.Timestamp.today()).strftime("%d/%m/%Y")
    logo_file = st.file_uploader("Logo opcional (PNG/JPG)", type=["png","jpg","jpeg"], key="logo_excel")

    if fuente == "Leer directamente de Ninox (tabla BASE DE DATOS)":
        try:
            with st.spinner("Leyendo registros desde Ninox…"):
                recs = ninox_list_records(TABLE_WRITE_NAME, limit=1000)
                df_nx = ninox_records_to_df(recs)
            if df_nx.empty:
                st.warning("No se recibieron registros desde Ninox.")
            else:
                st.session_state.df_final_vista = df_nx.copy()
                tmp = df_nx.copy()
                for h in ["Hp (10)","Hp (0.07)","Hp (3)"]:
                    tmp[h] = tmp[h].apply(hp_to_num)
                tmp["_Hp10_NUM"]  = tmp["Hp (10)"]; tmp["_Hp007_NUM"] = tmp["Hp (0.07)"]; tmp["_Hp3_NUM"] = tmp["Hp (3)"]
                st.session_state.df_final_num = tmp[["_Hp10_NUM","_Hp007_NUM","_Hp3_NUM","PERIODO DE LECTURA","CLIENTE",
                                                     "CÓDIGO DE USUARIO","CÓDIGO DE DOSÍMETRO","NOMBRE","CÉDULA",
                                                     "TIPO DE DOSÍMETRO","FECHA DE LECTURA"]].copy()
        except Exception as e:
            st.error(f"Error leyendo Ninox: {e}")

    df_vista = st.session_state.get("df_final_vista")
    df_num   = st.session_state.get("df_final_num")
    if df_vista is None or df_vista.empty or df_num is None or df_num.empty:
        st.info("No hay datos para mostrar en el reporte final.")
    else:
        clientes = sorted([c for c in df_vista["CLIENTE"].dropna().unique().tolist() if str(c).strip()])
        cliente_filtro = None
        if clientes:
            cliente_sel = st.selectbox("Filtrar por CLIENTE (opcional)", ["(Todos)"] + clientes, index=0)
            if cliente_sel != "(Todos)":
                cliente_filtro = cliente_sel
                df_vista = df_vista[df_vista["CLIENTE"] == cliente_filtro].copy()
                df_num   = df_num[df_num["CLIENTE"] == cliente_filtro].copy()

        reporte = construir_reporte_unico(df_vista, df_num, umbral_pm=0.005, agrupar_control_por="CLIENTE")
        if reporte.empty:
            st.info("No hay datos para el reporte con el filtro aplicado.")
        else:
            st.dataframe(reporte, use_container_width=True)

            # Descargas
            base = re.sub(r"[^A-Za-z0-9_\- ]+", "_", nombre_archivo_base.strip()) or "Reporte_Final"  # sanitize
            csv_bytes = reporte.to_csv(index=False).encode("utf-8-sig")
            st.download_button("⬇️ Descargar CSV (tabla)", data=csv_bytes, file_name=f"{base}.csv", mime="text/csv")

            excel_bytes = build_excel_like_example(
                df_reporte=reporte,
                fecha_emision=fecha_emision_ui,
                cliente=(cliente_filtro or "(Varios)"),
                codigo_reporte=codigo_reporte_ui or "SIN-CÓDIGO",
                logo_bytes=(logo_file.read() if logo_file else None),
            )
            st.download_button("⬇️ Descargar Excel (con diseño y notas)",
                               data=excel_bytes,
                               file_name=f"{base}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



